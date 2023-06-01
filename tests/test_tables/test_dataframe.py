# Copyright (C) 2023 Helmar Hutschenreuter
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.

from tempfile import NamedTemporaryFile

import pytest
from openpyxl import Workbook, load_workbook

from mvtool.tables.dataframe import Cell, DataFrame, read_excel, write_excel


def test_dataframe_init():
    df = DataFrame(
        [
            [Cell("A", 1), Cell("B", 2)],
            [Cell("A", 3), Cell("B", 4)],
        ]
    )
    assert df.data == {
        "A": [1, 3],
        "B": [2, 4],
    }


def test_dataframe_iter_rows():
    df = DataFrame(
        [
            [Cell("A", 1), Cell("B", 2)],
            [Cell("A", 3), Cell("B", 4)],
        ]
    )

    rows = [list(r) for r in list(df)]
    assert rows == [
        [Cell("A", 1), Cell("B", 2)],
        [Cell("A", 3), Cell("B", 4)],
    ]


def test_dataframe_row_count():
    df = DataFrame(
        [
            [Cell("A", 1), Cell("B", 2)],
            [Cell("A", 3), Cell("B", 4)],
            [Cell("A", 5), Cell("B", 6)],
        ]
    )
    assert len(df) == 3


def test_dataframe_row_count_empty():
    df = DataFrame()
    assert len(df) == 0


def test_dataframe_missing_value():
    df = DataFrame(
        [
            [Cell("A", 1), Cell("B", 2)],
            [Cell("A", 3)],
        ]
    )
    assert df.data == {
        "A": [1, 3],
        "B": [2, None],
    }


def test_dataframe_column_names():
    df = DataFrame(
        [
            [Cell("A", 1), Cell("B", 2)],
            [Cell("A", 3), Cell("B", 4)],
        ]
    )
    assert df.column_names == ["A", "B"]


def test_dataframe_getitem():
    df = DataFrame(
        [
            [Cell("Alpha", 1), Cell("Beta", 2)],
            [Cell("Alpha", 3), Cell("Beta", 4)],
        ]
    )
    sub_df = df["Alpha"]

    assert sub_df.data == {"Alpha": [1, 3]}
    assert sub_df.column_names == ["Alpha"]


def test_dataframe_getitem_multiple_columns():
    df = DataFrame(
        [
            [Cell("A", 1), Cell("B", 2), Cell("C", 3)],
            [Cell("A", 4), Cell("B", 5), Cell("C", 6)],
        ]
    )

    sub_df = df[["A", "B"]]
    assert sub_df.data == {
        "A": [1, 4],
        "B": [2, 5],
    }
    assert sub_df.column_names == ["A", "B"]


def test_dataframe_getitem_invalid_column():
    df = DataFrame([[Cell("A", 1), Cell("B", 2)], [Cell("A", 3), Cell("B", 4)]])

    with pytest.raises(KeyError):
        df["C"]


def test_read_excel():
    # Create a workbook and add some data
    wb = Workbook()
    ws = wb.active
    ws.append(["A", "B", "C"])
    ws.append([1, 2])
    ws.append([3, 4, 5])

    # Save the workbook to a temporary file
    with NamedTemporaryFile(suffix=".xlsx") as temp:
        wb.save(temp.name)

        # Now read the file with read_excel and check the data
        df = read_excel(temp.name)
        assert df.data == {
            "A": [1, 3],
            "B": [2, 4],
            "C": [None, 5],
        }


def test_write_excel():
    # Create a DataFrame and add some data
    df = DataFrame(
        [
            [Cell("A", 1), Cell("B", 2)],
            [Cell("A", 3), Cell("B", 4)],
        ]
    )

    # Save the DataFrame to a temporary file
    with NamedTemporaryFile(suffix=".xlsx") as temp:
        write_excel(df, temp.name)

        # Now read the file with openpyxl and check the data
        wb = load_workbook(temp.name)
        ws = wb.active

        assert ws["A1"].value == "A"
        assert ws["B1"].value == "B"
        assert ws["A2"].value == 1
        assert ws["B2"].value == 2
        assert ws["A3"].value == 3
        assert ws["B3"].value == 4


def test_write_excel_empty():
    # Create an empty DataFrame
    df = DataFrame([])

    # Save the DataFrame to a temporary file
    with NamedTemporaryFile(suffix=".xlsx") as temp:
        write_excel(df, temp.name)

        # Now read the file with openpyxl and check the data
        wb = load_workbook(temp.name)
        ws = wb.active

        # The worksheet should be empty
        assert ws.calculate_dimension() == "A1:A1"
