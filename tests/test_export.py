# coding: utf-8
# 
# Copyright 2022 Helmar Hutschenreuter
#
# The source code of this program is made available
# under the terms of the GNU Affero General Public License version 3
# (GNU AGPL V3) as published by the Free Software Foundation. You may obtain
# a copy of the GNU AGPL V3 at https://www.gnu.org/licenses/.
#
# In the case you use this program under the terms of the GNU AGPL V3,
# the program is provided in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
# GNU AGPL V3 for more details.

from unittest.mock import Mock
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from pydantic import ValidationError
import pytest

from mvtool.models import Measure, Project, Requirement, RequirementInput
from mvtool.views.export import ExportMeasuresView, ExportRequirementsView
from mvtool.views.import_ import ImportRequirementsView
from mvtool.views.requirements import RequirementsView

def test_query_measure_data(
        export_measures_view: ExportMeasuresView, create_project: Project, 
        create_measure: Measure):
    results = list(export_measures_view.query_measure_data(create_project.id))

    assert len(results) == 1
    result = results[0]
    assert isinstance(result, tuple)
    measure, requirement, document, jira_issue = result
    assert isinstance(measure, Measure)
    assert isinstance(requirement, Requirement)
    assert document == create_measure.document
    assert jira_issue == None

def test_download_measures_excel(
        export_measures_view: ExportMeasuresView, excel_temp_file, 
        create_project: Project, create_measure: Measure):
    result = export_measures_view.download_measures_excel(
        create_project.id, temp_file=excel_temp_file)
    assert isinstance(result, FileResponse)
    assert result.media_type == \
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

def test_download_requirements_excel(
        export_requirements_view: ExportRequirementsView, excel_temp_file,
        create_project: Project, create_requirement: Requirement):
    result = export_requirements_view.download_requirements_excel(
        create_project.id, temp_file=excel_temp_file)
    assert isinstance(result, FileResponse)
    assert result.media_type == \
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def test_read_requirements_from_excel_worksheet():
    sut = ImportRequirementsView(None, None)
    workbook = load_workbook('tests/import/valid.xlsx')
    worksheet = workbook.active

    results = list(sut.read_requirement_from_excel_worksheet(worksheet))

    assert len(results) >= 1
    result = results[0]
    assert isinstance(result, RequirementInput)

def test_read_requirements_from_excel_worksheet_invalid_headers():
    sut = ImportRequirementsView(None, None)
    workbook = load_workbook('tests/import/invalid_headers.xlsx')
    worksheet = workbook.active

    with pytest.raises(ValueError):
        list(sut.read_requirement_from_excel_worksheet(worksheet))

def test_read_requirements_from_excel_worksheet_invalid_data():
    sut = ImportRequirementsView(None, None)
    workbook = load_workbook('tests/import/invalid_data.xlsx')
    worksheet = workbook.active

    with pytest.raises(ValidationError):
        list(sut.read_requirement_from_excel_worksheet(worksheet))