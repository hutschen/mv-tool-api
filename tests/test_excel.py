# coding: utf-8
#
# Copyright (C) 2022 Helmar Hutschenreuter
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.

import io
from unittest.mock import Mock
from fastapi import HTTPException
from fastapi.responses import FileResponse
import pytest

from mvtool.models import (
    Document,
    Measure,
    Project,
    Requirement,
    RequirementInput,
    RequirementOutput,
)
from mvtool.views.excel import (
    DocumentsExcelView,
    ExcelHeader,
    ExcelView,
    MeasuresExcelView,
    RequirementsExcelView,
)
from openpyxl import Workbook


@pytest.fixture
def worksheet_rows():
    return [
        {"int": 0, "str": "hello", "bool": True, "float": 1.0},
        {"int": 1, "str": "world", "bool": False, "float": 2.0},
        {"int": 2, "str": None, "bool": False, "float": 2.5},
    ]


@pytest.fixture
def worksheet_headers():
    return [
        ExcelHeader("int"),
        ExcelHeader("str"),
        ExcelHeader("bool"),
        ExcelHeader("float"),
    ]


@pytest.fixture
def empty_worksheet():
    return Workbook().active


@pytest.fixture
def filled_worksheet(empty_worksheet, worksheet_rows):
    headers = None
    for row in worksheet_rows:
        if not headers:
            headers = [k for k, _ in row.items()]
            empty_worksheet.append(headers)
        empty_worksheet.append(row[h] for h in headers)
    return empty_worksheet


def test_read_worksheet(filled_worksheet, worksheet_headers, worksheet_rows):
    sut = ExcelView(worksheet_headers)
    sut._convert_from_row = lambda row, *_: row

    results = list(sut._read_worksheet(filled_worksheet))
    assert results == worksheet_rows


def test_read_worksheet_invalid_headers(filled_worksheet):
    sut = ExcelView([ExcelHeader("not_existing")])

    with pytest.raises(HTTPException) as error_info:
        list(sut._read_worksheet(filled_worksheet))

    assert error_info.value.status_code == 400
    assert error_info.value.detail.startswith("Missing headers")


def test_write_worksheet(empty_worksheet, worksheet_headers, worksheet_rows):
    sut = ExcelView(worksheet_headers)
    sut._convert_to_row = lambda row, *_: row

    sut._write_worksheet(empty_worksheet, worksheet_rows)

    # read back the worksheet and compare the rows
    headers = None
    results = []
    for values in empty_worksheet.iter_rows(values_only=True):
        if not headers:
            headers = values
        else:
            results.append(dict(zip(headers, values)))
    assert results == worksheet_rows


def test_write_worksheet_no_rows(empty_worksheet, worksheet_headers):
    sut = ExcelView(worksheet_headers)
    sut._write_worksheet(empty_worksheet, [])

    # read back the worksheet
    headers = None
    row_count = 0
    for values in empty_worksheet.iter_rows(values_only=True):
        if not headers:
            headers = values
        else:
            row_count += 1
    assert headers == tuple(h.name for h in worksheet_headers)
    assert row_count == 0


def test_determine_headers_to_write(empty_worksheet, worksheet_headers):
    # Set worksheet headers optional
    for h in worksheet_headers:
        h.optional = True

    sut = ExcelView(worksheet_headers)
    sut._convert_to_row = lambda row, *_: row

    # Write data to worksheet
    sut._write_worksheet(
        empty_worksheet,
        [
            {"int": 0, "str": "hello", "bool": True, "float": 0.0},
            {"int": 0, "str": "world", "bool": False, "float": 0.0},
            {"int": 0, "str": None, "bool": False, "float": 0.0},
        ],
    )

    # Read headers from worksheet
    for headers in empty_worksheet.iter_rows(values_only=True):
        break
    assert headers == ("str", "bool")


def test_query_measure_data(
    measures_excel_view: MeasuresExcelView,
    create_project: Project,
    create_measure: Measure,
):
    results = list(
        measures_excel_view._query_measure_data(
            Requirement.project_id == create_project.id
        )
    )

    assert len(results) == 1
    result = results[0]
    assert isinstance(result, tuple)
    measure, requirement, document, jira_issue = result
    assert isinstance(measure, Measure)
    assert isinstance(requirement, Requirement)
    assert document == create_measure.document
    assert jira_issue == None


def test_download_measures_excel_for_project(
    measures_excel_view: MeasuresExcelView,
    excel_temp_file,
    create_project: Project,
    create_measure: Measure,
):
    result = measures_excel_view.download_measures_excel_for_project(
        create_project.id, temp_file=excel_temp_file
    )
    assert isinstance(result, FileResponse)
    assert (
        result.media_type
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def test_download_measures_excel_for_requirement(
    measures_excel_view: MeasuresExcelView,
    excel_temp_file,
    create_requirement: Requirement,
    create_measure: Measure,
):
    result = measures_excel_view.download_measures_excel_for_requirement(
        create_requirement.id, temp_file=excel_temp_file
    )
    assert isinstance(result, FileResponse)
    assert (
        result.media_type
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def test_upload_measures_excel(
    measures_excel_view: MeasuresExcelView,
    excel_temp_file,
    create_requirement: Requirement,
):
    upload_file = Mock()
    upload_file.file = io.FileIO("tests/data/excel/measures_valid.xlsx", "r")

    list(
        measures_excel_view.upload_measures_excel(
            create_requirement.id, upload_file, excel_temp_file
        )
    )

    assert create_requirement.measures is not None
    assert len(create_requirement.measures) > 0


def test_download_requirements_excel(
    requirements_excel_view: RequirementsExcelView,
    excel_temp_file,
    create_project: Project,
    create_requirement: Requirement,
):
    result = requirements_excel_view.download_requirements_excel(
        create_project.id, temp_file=excel_temp_file
    )
    assert isinstance(result, FileResponse)
    assert (
        result.media_type
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def test_bulk_create_update_requirements(
    requirements_excel_view: RequirementsExcelView,
    create_project: Project,
    create_requirement: Requirement,
):
    data = [
        (create_requirement.id, RequirementInput(summary="update")),
        (None, RequirementInput(summary="create")),
    ]

    results = requirements_excel_view._bulk_create_update_requirements(
        create_project.id, data
    )

    assert len(results) == 2
    assert isinstance(results[0], RequirementOutput)
    assert results[0].summary == "update"
    assert isinstance(results[1], RequirementOutput)
    assert results[1].summary == "create"


def test_upload_requirements_excel(
    requirements_excel_view: RequirementsExcelView,
    excel_temp_file,
    create_project: Project,
):
    upload_file = Mock()
    upload_file.file = io.FileIO("tests/data/excel/requirements_valid.xlsx", "r")

    list(
        requirements_excel_view.upload_requirements_excel(
            create_project.id, upload_file, excel_temp_file
        )
    )

    assert create_project.requirements is not None
    assert len(create_project.requirements) > 0


def test_download_documents_excel(
    documents_excel_view: DocumentsExcelView,
    excel_temp_file,
    create_project: Project,
    create_document: Document,
):
    result = documents_excel_view.download_documents_excel(
        create_project.id, temp_file=excel_temp_file
    )
    assert isinstance(result, FileResponse)
    assert (
        result.media_type
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def test_upload_documents_excel(
    documents_excel_view: DocumentsExcelView,
    excel_temp_file,
    create_project: Project,
):
    upload_file = Mock()
    upload_file.file = io.FileIO("tests/data/excel/documents_valid.xlsx", "r")

    list(
        documents_excel_view.upload_documents_excel(
            create_project.id, upload_file, excel_temp_file
        )
    )

    assert create_project.documents is not None
    assert len(create_project.documents) > 0
