# Copyright (C) 2023 Helmar Hutschenreuter
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.

from typing import IO, Any, Iterable, NamedTuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

from ..utils.errors import ValueHttpError


class Cell(NamedTuple):
    label: str
    value: Any


Row = Iterable[Cell]


class DataFrame:
    def __init__(self, rows: Iterable[Row] = None):
        rows = rows or []
        self.data = {}

        for row in rows:
            column_names = set(self.data.keys())
            labels = set()

            # Fill in values
            for chell in row:
                try:
                    self.data[chell.label].append(chell.value)
                except KeyError:
                    self.data[chell.label] = [chell.value]
                labels.add(chell.label)

            # Fill in None for missing values
            for column_name in column_names - labels:
                self.data[column_name].append(None)

    def __iter__(self):
        for values in zip(*self.data.values()):
            yield (Cell(l, v) for l, v in zip(self.data.keys(), values))

    def __len__(self):
        for _, values in self.data.items():
            return len(values)
        return 0

    @property
    def column_names(self) -> list[str]:
        return list(self.data.keys())

    def __getitem__(self, column_names: Iterable[str] | str) -> list[Any]:
        column_names = [column_names] if isinstance(column_names, str) else column_names
        df = DataFrame()
        df.data = {column_name: self.data[column_name] for column_name in column_names}
        return df


def _iter_rows(worksheet: Worksheet):
    for i, values in enumerate(worksheet.iter_rows(values_only=True)):
        if i == 0:
            # First row is header row
            labels = tuple(values)
        else:
            # Convert row to dict and yield
            yield (Cell(l, v) for l, v in zip(labels, values))


def read_excel(file_obj: str | IO[bytes]) -> DataFrame:
    # carefully open the Excel file
    try:
        workbook = load_workbook(file_obj, read_only=True)
    except Exception:
        # have to catch all exceptions, because openpyxl does raise several
        # exceptions when reading an invalid Excel file
        raise ValueHttpError("Excel file seems to be corrupted")

    # Load data from workbook
    return DataFrame(_iter_rows(workbook.active))


def write_excel(
    df: DataFrame, file_obj: str | IO[bytes], sheet_name: str | None = None
):
    # Create workbook
    workbook = Workbook()
    worksheet: Worksheet = workbook.active
    if sheet_name:
        worksheet.title = sheet_name

    # Fill worksheet with data
    worksheet.append(df.column_names)
    is_empty = True
    for row in df:
        worksheet.append(cell.value for cell in row)
        is_empty = False

    # Add table to worksheet
    if not is_empty:
        table = Table(displayName=worksheet.title, ref=worksheet.calculate_dimension())
        worksheet.add_table(table)

    workbook.save(file_obj)
