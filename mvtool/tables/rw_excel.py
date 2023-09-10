# Copyright (C) 2023 Helmar Hutschenreuter
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.

import re
from typing import IO

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

from ..utils.errors import ValueHttpError
from .dataframe import Cell, DataFrame


def _iter_rows(worksheet: Worksheet):
    for i, values in enumerate(worksheet.iter_rows(values_only=True)):
        if i == 0:
            # First row is header row
            labels = tuple(values)
        else:
            # Convert row to dict and yield
            yield (Cell(l, v) for l, v in zip(labels, values))


def read_excel(file_obj: str | IO[bytes]) -> DataFrame:
    # carefully open the Excel file
    try:
        workbook = load_workbook(file_obj, read_only=True)
    except Exception:
        # have to catch all exceptions, because openpyxl does raise several
        # exceptions when reading an invalid Excel file
        raise ValueHttpError("Excel file seems to be corrupted")

    # Load data from workbook
    return DataFrame(_iter_rows(workbook.active))


def write_excel(
    df: DataFrame, file_obj: str | IO[bytes], sheet_name: str | None = None
):
    # Create workbook
    workbook = Workbook()
    worksheet: Worksheet = workbook.active
    if sheet_name:
        worksheet.title = sheet_name

    # Fill worksheet with data
    worksheet.append(df.column_names)
    is_empty = True
    for row in df:
        worksheet.append(cell.value for cell in row)
        is_empty = False

    # Add table to worksheet
    if not is_empty:
        # Create table name from sheet name
        table_name = worksheet.title.lower()
        table_name = re.sub(r"[^a-z0-9]", "_", table_name)

        table = Table(displayName=table_name, ref=worksheet.calculate_dimension())
        worksheet.add_table(table)

    workbook.save(file_obj)
