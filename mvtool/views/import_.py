# coding: utf-8
# 
# Copyright 2022 Helmar Hutschenreuter
#
# The source code of this program is made available
# under the terms of the GNU Affero General Public License version 3
# (GNU AGPL V3) as published by the Free Software Foundation. You may obtain
# a copy of the GNU AGPL V3 at https://www.gnu.org/licenses/.
#
# In the case you use this program under the terms of the GNU AGPL V3,
# the program is provided in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
# GNU AGPL V3 for more details.

from tempfile import NamedTemporaryFile
from fastapi import APIRouter, Depends, Response, UploadFile
from fastapi_utils.cbv import cbv
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
from pydantic import ValidationError
from mvtool.models import MeasureInput, RequirementInput
from mvtool.views.export import get_excel_temp_file

from mvtool.views.requirements import RequirementsView
from mvtool.views.measures import MeasuresView
from mvtool import errors

router = APIRouter()

@cbv(router)
class ImportRequirementsView:
    kwargs = dict(tags=['requirement'])

    def __init__(self,
            temp_file: NamedTemporaryFile = Depends(get_excel_temp_file),
            requirements: RequirementsView = Depends(RequirementsView)):
        self._temp_file = temp_file
        self._requirements = requirements

    def read_requirement_from_excel_worksheet(self, worksheet: Worksheet):
        headers = {
            'Reference', 'Summary', 'Description', 'Target Object', 
            'Compliance Status', 'Compliance Comment'}
        
        for index, values in enumerate(worksheet.iter_rows(values_only=True)):
            # check and process header row
            if index == 0:
                # check if all required headers are present
                if not headers.issubset(values):
                    detail = 'Missing headers on worksheet "%s": %s' % (
                        worksheet.title, ', '.join(headers - set(values)))
                    raise errors.ValueHttpError(detail)
                headers = tuple(values)
                continue
            
            # get data from row
            requirement_data = dict(zip(headers, values))
            try:
                requirement_input = RequirementInput(
                    reference=requirement_data['Reference'] or None,
                    summary=requirement_data['Summary'],
                    description=requirement_data['Description'] or None,
                    target_object=requirement_data['Target Object'] or None,
                    compliance_status=requirement_data['Compliance Status'] or None,
                    compliance_comment=requirement_data['Compliance Comment'] or None)
            except ValidationError as error:
                detail = 'Invalid data on worksheet "%s" at row %d: %s' % (
                    worksheet.title, index + 1, error)
                raise errors.ValueHttpError(detail)
            else:
                yield requirement_input

    @router.post(
        '/projects/{project_id}/requirements/excel', status_code=201, 
        response_class=Response, **kwargs)
    def upload_requirements_excel(
            self, project_id: int, excel_file: UploadFile):
        with open(self._temp_file.name, 'wb') as f:
            # 1MB buffer size should be sufficient to load an Excel file
            buffer_size = 1000 * 1024
            chunk = excel_file.file.read(buffer_size)
            while chunk:
                f.write(chunk)
                chunk = excel_file.file.read(buffer_size)

        # carefully open the Excel file
        try:
            workbook = load_workbook(self._temp_file.name, read_only=True)
        except Exception:
            # have to catch all exceptions, because openpyxl does raise several
            # exceptions when reading an invalid Excel file
            raise errors.ValueHttpError('Excel file seems to be corrupt')
        worksheet = workbook.active
        if not worksheet:
            # when the Excel file is empty or not exists, openpyxl returns None 
            # instead of raising an exception
            raise errors.ValueHttpError('No worksheet found in Excel file')

        # read data from worksheet
        for requirement_input in \
                self.read_requirement_from_excel_worksheet(worksheet):
            self._requirements.create_requirement(
                project_id, requirement_input)


class ImportMeasuresView:
    kwargs = dict(tags=['measure'])

    def __init__(self,
            temp_file: NamedTemporaryFile = Depends(get_excel_temp_file),
            measures: MeasuresView = Depends(MeasuresView)):
        self._temp_file = temp_file
        self._measures = measures

    def read_measures_from_excel_worksheet(self, worksheet: Worksheet):
        headers = {'Summary', 'Description'}

        for index, values in enumerate(worksheet.iter_rows(values_only=True)):
            # check and process header row
            if index == 0:
                # check if all required headers are present
                if not headers.issubset(values):
                    detail = 'Missing headers on worksheet "%s": %s' % (
                        worksheet.title, ', '.join(headers - set(values)))
                    raise errors.ValueHttpError(detail)
                headers = tuple(values)
                continue
            
            # get data from row
            measure_data = dict(zip(headers, values))
            try:
                measure_input = MeasureInput(
                    summary=measure_data['Summary'],
                    description=measure_data['Description'] or None)
            except ValidationError as error:
                detail = 'Invalid data on worksheet "%s" at row %d: %s' % (
                    worksheet.title, index + 1, error)
                raise errors.ValueHttpError(detail)
            else:
                yield measure_input

    @router.post(
        '/requirements/{requirement_id}/measures/excel', status_code=201,
        response_class=Response, **kwargs)
    def upload_measures_excel(
            self, requirement_id: int, excel_file: UploadFile):
        with open(self._temp_file.name, 'wb') as f:
            # 1MB buffer size should be sufficient to load an Excel file
            buffer_size = 1000 * 1024
            chunk = excel_file.file.read(buffer_size)
            while chunk:
                f.write(chunk)
                chunk = excel_file.file.read(buffer_size)

        # carefully open the Excel file
        try:
            workbook = load_workbook(self._temp_file.name, read_only=True)
        except Exception:
            # have to catch all exceptions, because openpyxl does raise several
            # exceptions when reading an invalid Excel file
            raise errors.ValueHttpError('Excel file seems to be corrupt')
        worksheet = workbook.active
        if not worksheet:
            # when the Excel file is empty or not exists, openpyxl returns None 
            # instead of raising an exception
            raise errors.ValueHttpError('No worksheet found in Excel file')

        # read data from worksheet
        for measure_input in \
                self.read_measures_from_excel_worksheet(worksheet):
            self._measures.create_measure(requirement_id, measure_input)
    
    @router.post(
        '/requirements/{requirement_id}/measures/excel', status_code=201,
        response_class=Response, **kwargs)
    def upload_measures_excel(
            self, requirement_id: int, excel_file: UploadFile):
        with open(self._temp_file.name, 'wb') as f:
            # 1MB buffer size should be sufficient to load an Excel file
            buffer_size = 1000 * 1024
            chunk = excel_file.file.read(buffer_size)
            while chunk:
                f.write(chunk)
                chunk = excel_file.file.read(buffer_size)

        # carefully open the Excel file
        try:
            workbook = load_workbook(self._temp_file.name, read_only=True)
        except Exception:
            # have to catch all exceptions, because openpyxl does raise several
            # exceptions when reading an invalid Excel file
            raise errors.ValueHttpError('Excel file seems to be corrupt')
        worksheet = workbook.active
        if not worksheet:
            # when the Excel file is empty or not exists, openpyxl returns None 
            # instead of raising an exception
            raise errors.ValueHttpError('No worksheet found in Excel file')

        # read data from worksheet
        for measure_input in \
                self.read_measures_from_excel_worksheet(worksheet):
            self._measures.create_measure(requirement_id, measure_input)