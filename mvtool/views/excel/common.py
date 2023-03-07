# coding: utf-8
#
# Copyright (C) 2023 Helmar Hutschenreuter
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.

from collections import OrderedDict
import shutil
from tempfile import NamedTemporaryFile
from typing import Collection, Generic, Iterator, TypeVar
from fastapi import UploadFile
from fastapi.responses import FileResponse
from pydantic import BaseModel, constr
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table

from ... import errors


def get_excel_temp_file():
    # TODO: Make this function more generic and move it to utils
    with NamedTemporaryFile(suffix=".xlsx") as temp_file:
        yield temp_file


class IdModel(BaseModel):
    # TODO: Move this to models
    id: int | None


class JiraIssueKeyModel(BaseModel):
    # TODO: Move this to models
    key: constr(regex=r"^[A-Za-z0-9\-]+$") | None


class ExcelHeader:
    READ_WRITE = 0
    READ_ONLY = 1
    WRITE_ONLY = 2

    def __init__(self, name: str, mode: int | None = None, optional: bool = False):
        self.name = name
        self._mode = mode or self.READ_WRITE
        self.optional = optional

    @property
    def is_write(self) -> bool:
        return self._mode in (self.READ_WRITE, self.WRITE_ONLY)

    @property
    def is_read(self) -> bool:
        return self._mode in (self.READ_WRITE, self.READ_ONLY)


T = TypeVar("T")


class ExcelView(Generic[T]):
    def __init__(self, headers: Collection[ExcelHeader]):
        self._write_headers = [header for header in headers if header.is_write]
        self._read_headers = [header for header in headers if header.is_read]

    def _convert_to_row(self, data: T, *args) -> dict[str, str]:
        raise NotImplementedError("Must be implemented by subclass")

    def _write_worksheet(
        self,
        worksheet: Worksheet,
        data: Iterator[T],
        *args,
    ):
        # Convert data to rows and determine optional headers
        header_flags = OrderedDict(
            (h.name, not h.optional) for h in self._write_headers
        )
        rows = []

        for row_data in data:
            row = self._convert_to_row(row_data, *args)
            for header_name, header_flag in header_flags.items():
                if not header_flag and not row[header_name]:
                    continue
                header_flags[header_name] = True
            rows.append(row)

        # Fill worksheet with data
        header_names = [h_name for h_name, h_flag in header_flags.items() if h_flag]
        worksheet.append(header_names)
        is_empty = True
        for row in rows:
            values = [row.get(h_name, "") for h_name in header_names]
            worksheet.append(values)
            is_empty = False

        # Add table to worksheet
        if not is_empty:
            table = Table(
                displayName=worksheet.title, ref=worksheet.calculate_dimension()
            )
            worksheet.add_table(table)

    def _convert_from_row(self, row: dict[str, str], worksheet, row_no, *args) -> T:
        raise NotImplementedError("Must be implemented by subclass")

    def _read_worksheet(self, worksheet: Worksheet, *args) -> Iterator[T]:
        required_header_names = {h.name for h in self._read_headers if not h.optional}
        worksheet_header_names = None
        is_header_row = True

        for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
            if is_header_row:
                # Check if all headers are present
                if not required_header_names.issubset(row):
                    detail = 'Missing headers on worksheet "%s": %s' % (
                        worksheet.title,
                        ", ".join(required_header_names - set(row)),
                    )
                    raise errors.ValueHttpError(detail)
                worksheet_header_names = tuple(row)
                is_header_row = False
            else:
                # Convert row to dict and yield
                yield self._convert_from_row(
                    {
                        **{h.name: None for h in self._read_headers},
                        **dict(zip(worksheet_header_names, row)),
                    },
                    worksheet,
                    row_index + 1,
                    *args,
                )

    def _process_download(
        self,
        data: Iterator[T],
        temp_file: NamedTemporaryFile,
        *args,
        sheet_name: str = "Export",
        filename: str = "export.xlsx",
    ) -> FileResponse:
        # set up workbook
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name

        # Write data to worksheet, save workbook and return file response
        self._write_worksheet(worksheet, data, *args)
        workbook.save(temp_file.name)
        return FileResponse(temp_file.name, filename=filename)

    def _process_upload(
        self, upload_file: UploadFile, temp_file: NamedTemporaryFile, *args
    ) -> Iterator[T]:
        # Save uploaded file to temp file
        shutil.copyfileobj(upload_file.file, temp_file.file)

        # carefully open the Excel file
        try:
            workbook = load_workbook(temp_file.name, read_only=True)
        except Exception:
            # have to catch all exceptions, because openpyxl does raise several
            # exceptions when reading an invalid Excel file
            raise errors.ValueHttpError("Excel file seems to be corrupted")

        # Load data from workbook
        worksheet = workbook.active
        return self._read_worksheet(worksheet, *args)
