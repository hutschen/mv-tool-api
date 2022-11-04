# coding: utf-8
#
# Copyright (C) 2022 Helmar Hutschenreuter
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.


from collections import OrderedDict
import shutil
from tempfile import NamedTemporaryFile
from typing import Collection, Generic, TypeVar
from fastapi import APIRouter, Depends, UploadFile
from fastapi.responses import FileResponse
from fastapi_utils.cbv import cbv
from pydantic import ValidationError, BaseModel, constr
from pyparsing import Iterator
from sqlmodel import Session, select
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table

from mvtool import errors
from mvtool.database import get_session
from mvtool.models import (
    Document,
    DocumentOutput,
    JiraIssue,
    Measure,
    MeasureInput,
    MeasureOutput,
    Requirement,
    RequirementInput,
    DocumentInput,
    RequirementOutput,
)
from .jira_ import JiraIssuesView
from .measures import MeasuresView
from .requirements import RequirementsView
from .documents import DocumentsView
from .projects import ProjectsView


def get_excel_temp_file():
    with NamedTemporaryFile(suffix=".xlsx") as temp_file:
        yield temp_file


router = APIRouter()
T = TypeVar("T")


class IdModel(BaseModel):
    id: int | None


class JiraIssueKeyModel(BaseModel):
    key: constr(regex=r"^[A-Za-z0-9\-]+$") | None


class ExcelHeader:
    READ_WRITE = 0
    READ_ONLY = 1
    WRITE_ONLY = 2

    def __init__(self, name: str, mode: int | None = None, optional: bool = False):
        self.name = name
        self._mode = mode or self.READ_WRITE
        self.optional = optional

    @property
    def is_write(self) -> bool:
        return self._mode in (self.READ_WRITE, self.WRITE_ONLY)

    @property
    def is_read(self) -> bool:
        return self._mode in (self.READ_WRITE, self.READ_ONLY)


class ExcelView(Generic[T]):
    def __init__(self, headers: Collection[ExcelHeader]):
        self._write_headers = [header for header in headers if header.is_write]
        self._read_headers = [header for header in headers if header.is_read]

    def _convert_to_row(self, data: T, *args) -> dict[str, str]:
        raise NotImplementedError("Must be implemented by subclass")

    def _write_worksheet(
        self,
        worksheet: Worksheet,
        data: Iterator[T],
        *args,
    ):
        # Convert data to rows and determine optional headers
        header_flags = OrderedDict(
            (h.name, not h.optional) for h in self._write_headers
        )
        rows = []

        for row_data in data:
            row = self._convert_to_row(row_data, *args)
            for header_name, header_flag in header_flags.items():
                if not header_flag and not row[header_name]:
                    continue
                header_flags[header_name] = True
            rows.append(row)

        # Fill worksheet with data
        header_names = [h_name for h_name, h_flag in header_flags.items() if h_flag]
        worksheet.append(header_names)
        is_empty = True
        for row in rows:
            values = [row.get(h_name, "") for h_name in header_names]
            worksheet.append(values)
            is_empty = False

        # Add table to worksheet
        if not is_empty:
            table = Table(
                displayName=worksheet.title, ref=worksheet.calculate_dimension()
            )
            worksheet.add_table(table)

    def _convert_from_row(self, row: dict[str, str], worksheet, row_no, *args) -> T:
        raise NotImplementedError("Must be implemented by subclass")

    def _read_worksheet(self, worksheet: Worksheet, *args) -> Iterator[T]:
        required_header_names = {h.name for h in self._read_headers if not h.optional}
        worksheet_header_names = None
        is_header_row = True

        for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
            if is_header_row:
                # Check if all headers are present
                if not required_header_names.issubset(row):
                    detail = 'Missing headers on worksheet "%s": %s' % (
                        worksheet.title,
                        ", ".join(required_header_names - set(row)),
                    )
                    raise errors.ValueHttpError(detail)
                worksheet_header_names = tuple(row)
                is_header_row = False
            else:
                # Convert row to dict and yield
                yield self._convert_from_row(
                    {
                        **{h.name: None for h in self._read_headers},
                        **dict(zip(worksheet_header_names, row)),
                    },
                    worksheet,
                    row_index + 1,
                    *args,
                )

    def _process_download(
        self,
        data: Iterator[T],
        temp_file: NamedTemporaryFile,
        *args,
        sheet_name: str = "Export",
        filename: str = "export.xlsx",
    ) -> FileResponse:
        # set up workbook
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name

        # Write data to worksheet, save workbook and return file response
        self._write_worksheet(worksheet, data, *args)
        workbook.save(temp_file.name)
        return FileResponse(temp_file.name, filename=filename)

    def _process_upload(
        self, upload_file: UploadFile, temp_file: NamedTemporaryFile, *args
    ) -> Iterator[T]:
        # Save uploaded file to temp file
        shutil.copyfileobj(upload_file.file, temp_file.file)

        # carefully open the Excel file
        try:
            workbook = load_workbook(temp_file.name, read_only=True)
        except Exception:
            # have to catch all exceptions, because openpyxl does raise several
            # exceptions when reading an invalid Excel file
            raise errors.ValueHttpError("Excel file seems to be corrupted")

        # Load data from workbook
        worksheet = workbook.active
        return self._read_worksheet(worksheet, *args)


@cbv(router)
class MeasuresExcelView(ExcelView):
    kwargs = MeasuresView.kwargs

    def __init__(
        self,
        session: Session = Depends(get_session),
        jira_issues: JiraIssuesView = Depends(JiraIssuesView),
        measures: MeasuresView = Depends(MeasuresView),
    ):
        ExcelView.__init__(
            self,
            [
                ExcelHeader("Requirement Reference", ExcelHeader.WRITE_ONLY, True),
                ExcelHeader("Requirement GS ID", ExcelHeader.WRITE_ONLY, True),
                ExcelHeader("Requirement Summary", ExcelHeader.WRITE_ONLY, True),
                ExcelHeader("ID", optional=True),
                ExcelHeader("Summary"),
                ExcelHeader("Description", optional=True),
                ExcelHeader("Completed", optional=True),
                ExcelHeader("Document Reference", ExcelHeader.WRITE_ONLY, True),
                ExcelHeader("Document Title", ExcelHeader.WRITE_ONLY, True),
                ExcelHeader("JIRA Issue Key", optional=True),
            ],
        )
        self._session = session
        self._jira_issues = jira_issues
        self._measures = measures

    def _convert_to_row(self, data: Measure, *args) -> dict[str, str]:
        return {
            "Requirement Reference": data.requirement.reference,
            "Requirement GS ID": (
                data.requirement.catalog_requirement.gs_anforderung_reference
                if data.requirement.catalog_requirement
                else None
            ),
            "Requirement Summary": data.requirement.summary,
            "ID": data.id,
            "Summary": data.summary,
            "Description": data.description,
            "Completed": data.completed,
            "Document Reference": data.document.reference if data.document else None,
            "Document Title": data.document.title if data.document else None,
            "JIRA Issue Key": data.jira_issue.key if data.jira_issue else None,
        }

    @router.get(
        "/projects/{project_id}/measures/excel", response_class=FileResponse, **kwargs
    )
    def download_measures_excel_for_project(
        self,
        project_id: int,
        sheet_name: str = "Export",
        filename: str = "export.xlsx",
        temp_file: NamedTemporaryFile = Depends(get_excel_temp_file),
    ) -> FileResponse:
        return self._process_download(
            self._measures.list_measures_of_project(project_id),
            temp_file,
            sheet_name=sheet_name,
            filename=filename,
        )

    @router.get(
        "/requirements/{requirement_id}/measures/excel",
        response_class=FileResponse,
        **kwargs,
    )
    def download_measures_excel_for_requirement(
        self,
        requirement_id: int,
        sheet_name: str = "Export",
        filename: str = "export.xlsx",
        temp_file: NamedTemporaryFile = Depends(get_excel_temp_file),
    ) -> FileResponse:
        return self._process_download(
            self._measures.list_measures(requirement_id),
            temp_file,
            sheet_name=sheet_name,
            filename=filename,
        )

    def _convert_from_row(
        self, row: dict[str, str], worksheet, row_no: int
    ) -> tuple[int | None, str | None, MeasureInput]:
        try:
            measure_id = IdModel(id=row["ID"]).id
            jira_issue_key = JiraIssueKeyModel(key=row["JIRA Issue Key"]).key
            measure_input = MeasureInput(
                summary=row["Summary"],
                description=row["Description"] or None,
                completed=row["Completed"] or False,
            )
        except ValidationError as error:
            detail = 'Invalid data on worksheet "%s" at row %d: %s' % (
                worksheet.title,
                row_no + 1,
                error,
            )
            raise errors.ValueHttpError(detail)
        else:
            return measure_id, jira_issue_key, measure_input

    def _bulk_create_patch_measures(
        self,
        requirement_id: int,
        data: Iterator[tuple[int | None, str | None, MeasureInput]],
    ) -> list[Measure]:
        # TODO: Define this attribute in constructor
        self._requirements = self._measures._requirements
        self._documents = self._measures._documents

        # Get requirement from database and retrieve data
        requirement = self._requirements.get_requirement(requirement_id)
        data = list(data)

        # Retrieve jira issues to be linked to measures
        jira_issues = self._jira_issues.get_jira_issues(
            [ji_key for _, ji_key, _ in data if ji_key]
        )
        jira_issue_map = {ji.key: ji for ji in jira_issues}

        # Read measures to be updated from database
        query = select(Measure).where(
            Measure.id.in_({id for id, _, _ in data if id is not None}),
            Measure.requirement_id == requirement_id,
        )
        read_measures = dict((m.id, m) for m in self._session.exec(query).all())

        # Create or update measures
        written_measures = []
        for measure_id, jira_issue_key, measure_input in data:
            jira_issue = None

            # Set jira issue id
            if jira_issue_key:
                jira_issue = jira_issue_map.get(jira_issue_key)
                if jira_issue is None:
                    raise errors.NotFoundError(f"JIRA issue {jira_issue_key} not found")
                measure_input.jira_issue_id = jira_issue.id

            if measure_id is None:
                # Create measure
                measure = Measure.from_orm(measure_input)
                measure.requirement = requirement
            else:
                # Update measure
                measure = read_measures.get(measure_id)
                if measure is None:
                    raise errors.NotFoundError(
                        "Measure with ID %d not part of requirement with ID %d"
                        % (measure_id, requirement_id)
                    )
                for key, value in measure_input.dict(exclude_unset=True).items():
                    setattr(measure, key, value)

            # set jira project and issue
            self._measures._set_jira_project_and_issue(measure, jira_issue=jira_issue)

            # TODO: Checking document id should be done more efficiently
            self._documents.check_document_id(measure.document_id)
            self._session.add(measure)
            written_measures.append(measure)

        self._session.flush()
        return written_measures

    @router.post(
        "/requirements/{requirement_id}/measures/excel",
        status_code=201,
        response_model=list[MeasureOutput],
        **kwargs,
    )
    def upload_measures_excel(
        self,
        requirement_id: int,
        upload_file: UploadFile,
        temp_file: NamedTemporaryFile = Depends(get_excel_temp_file),
    ) -> Iterator[MeasureOutput]:
        return self._bulk_create_patch_measures(
            requirement_id, self._process_upload(upload_file, temp_file)
        )


@cbv(router)
class RequirementsExcelView(ExcelView):
    kwargs = RequirementsView.kwargs

    def __init__(
        self,
        session: Session = Depends(get_session),
        projects: ProjectsView = Depends(ProjectsView),
        requirements: RequirementsView = Depends(RequirementsView),
    ):
        ExcelView.__init__(
            self,
            [
                ExcelHeader("ID", optional=True),
                ExcelHeader("Reference", optional=True),
                ExcelHeader("GS ID", ExcelHeader.WRITE_ONLY, True),
                ExcelHeader("Catalog Module", ExcelHeader.WRITE_ONLY, True),
                ExcelHeader("Summary"),
                ExcelHeader("Description", optional=True),
                ExcelHeader("GS Absicherung", ExcelHeader.WRITE_ONLY, True),
                ExcelHeader("GS Verantwortliche", ExcelHeader.WRITE_ONLY, True),
                ExcelHeader("Target Object", optional=True),
                ExcelHeader("Compliance Status", optional=True),
                ExcelHeader("Compliance Comment", optional=True),
                ExcelHeader("Completion", ExcelHeader.WRITE_ONLY, True),
            ],
        )
        self._session = session
        self._projects = projects
        self._requirements = requirements

    def _convert_to_row(self, data: Requirement) -> dict[str, str]:
        return {
            "ID": data.id,
            "Reference": data.reference,
            "GS ID": (
                data.catalog_requirement.gs_anforderung_reference
                if data.catalog_requirement
                else None
            ),
            "Catalog Module": (
                data.catalog_requirement.catalog_module.title
                if data.catalog_requirement
                else None
            ),
            "Summary": data.summary,
            "Description": data.description,
            "GS Absicherung": (
                data.catalog_requirement.gs_absicherung
                if data.catalog_requirement
                else None
            ),
            "GS Verantwortliche": (
                data.catalog_requirement.gs_verantwortliche
                if data.catalog_requirement
                else None
            ),
            "Target Object": data.target_object,
            "Compliance Status": data.compliance_status,
            "Compliance Comment": data.compliance_comment,
            "Completion": data.completion,
        }

    @router.get(
        "/projects/{project_id}/requirements/excel",
        response_class=FileResponse,
        **kwargs,
    )
    def download_requirements_excel(
        self,
        project_id: int,
        sheet_name: str = "Export",
        filename: str = "export.xlsx",
        temp_file: NamedTemporaryFile = Depends(get_excel_temp_file),
    ) -> FileResponse:
        return self._process_download(
            self._requirements.list_requirements(project_id),
            temp_file,
            sheet_name=sheet_name,
            filename=filename,
        )

    def _convert_from_row(
        self, row: dict[str, str], worksheet, row_no: int
    ) -> tuple[int | None, RequirementInput]:
        try:
            requirement_id = IdModel(id=row["ID"]).id
            requirement_input = RequirementInput(
                reference=row["Reference"] or None,
                summary=row["Summary"],
                description=row["Description"] or None,
                target_object=row["Target Object"] or None,
                compliance_status=row["Compliance Status"] or None,
                compliance_comment=row["Compliance Comment"] or None,
            )
        except ValidationError as error:
            detail = 'Invalid data on worksheet "%s" at row %d: %s' % (
                worksheet.title,
                row_no + 1,
                error,
            )
            raise errors.ValueHttpError(detail)
        else:
            return requirement_id, requirement_input

    def _bulk_create_update_requirements(
        self, project_id: int, data: Iterator[tuple[int | None, RequirementInput]]
    ) -> list[Requirement]:
        # Get project from database and retrieve data
        project = self._projects.get_project(project_id)
        data = list(data)

        # Read requirements to be updated from database
        query = select(Requirement).where(
            Requirement.id.in_({id for id, _ in data if id is not None}),
            Requirement.project_id == project_id,
        )
        read_requirements = dict((r.id, r) for r in self._session.exec(query).all())

        # Create or update requirements
        written_requirements = []
        for requirement_id, requirement_input in data:
            if requirement_id is None:
                # Create requirement
                requirement = Requirement.from_orm(requirement_input)
                requirement.project = project
            else:
                # Update requirement
                requirement = read_requirements.get(requirement_id)
                if requirement is None:
                    raise errors.NotFoundError(
                        "Requirement with ID %d not part of project with ID %d"
                        % (requirement_id, project_id)
                    )
                for key, value in requirement_input.dict().items():
                    setattr(requirement, key, value)

            self._requirements._set_jira_project(requirement)
            self._session.add(requirement)
            written_requirements.append(requirement)

        self._session.flush()
        return written_requirements

    @router.post(
        "/projects/{project_id}/requirements/excel",
        status_code=201,
        response_model=list[RequirementOutput],
        **kwargs,
    )
    def upload_requirements_excel(
        self,
        project_id: int,
        upload_file: UploadFile,
        temp_file: NamedTemporaryFile = Depends(get_excel_temp_file),
    ) -> Iterator[Requirement]:
        return self._bulk_create_update_requirements(
            project_id, self._process_upload(upload_file, temp_file)
        )


@cbv(router)
class DocumentsExcelView(ExcelView):
    kwargs = DocumentsView.kwargs

    def __init__(
        self,
        session: Session = Depends(get_session),
        projects: ProjectsView = Depends(ProjectsView),
        documents: DocumentsView = Depends(DocumentsView),
    ):
        ExcelView.__init__(
            self,
            [
                ExcelHeader("ID", optional=True),
                ExcelHeader("Reference", optional=True),
                ExcelHeader("Title"),
                ExcelHeader("Description", optional=True),
            ],
        )
        self._session = session
        self._projects = projects
        self._documents = documents

    def _convert_to_row(self, data: Document) -> dict[str, str]:
        return {
            "ID": data.id,
            "Reference": data.reference,
            "Title": data.title,
            "Description": data.description,
        }

    @router.get(
        "/projects/{project_id}/documents/excel",
        response_class=FileResponse,
        **kwargs,
    )
    def download_documents_excel(
        self,
        project_id: int,
        sheet_name: str = "Export",
        filename: str = "export.xlsx",
        temp_file: NamedTemporaryFile = Depends(get_excel_temp_file),
    ) -> FileResponse:
        return self._process_download(
            self._documents.list_documents(project_id),
            temp_file,
            sheet_name=sheet_name,
            filename=filename,
        )

    def _convert_from_row(
        self, row: dict[str, str], worksheet, row_no: int
    ) -> tuple[int | None, DocumentInput]:
        # Convert the row to a DocumentInput
        try:
            document_id = IdModel(id=row["ID"]).id
            document_input = DocumentInput(
                reference=row["Reference"] or None,
                title=row["Title"],
                description=row["Description"] or None,
            )
        except ValidationError as error:
            detail = 'Invalid data on worksheet "%s" at row %d: %s' % (
                worksheet.title,
                row_no + 1,
                error,
            )
            raise errors.ValueHttpError(detail)
        else:
            return document_id, document_input

    def _bulk_create_update_documents(
        self, project_id: int, data: Iterator[tuple[int | None, DocumentInput]]
    ) -> list[Document]:
        # Get project from database and retrieve data
        project = self._projects.get_project(project_id)
        data = list(data)

        # Read documents to be updated from database
        query = select(Document).where(
            Document.id.in_({id for id, _ in data if id is not None}),
            Document.project_id == project_id,
        )
        read_documents = dict((r.id, r) for r in self._session.exec(query).all())

        # Create or update documents
        written_documents = []
        for document_id, document_input in data:
            if document_id is None:
                # Create document
                document = Document.from_orm(document_input)
                document.project = project
            else:
                # Update document
                document = read_documents.get(document_id)
                if document is None:
                    raise errors.NotFoundError(
                        "Document with ID %d not part of project with ID %d"
                        % (document_id, project_id)
                    )
                for key, value in document_input.dict().items():
                    setattr(document, key, value)

            self._documents._set_jira_project(document)
            self._session.add(document)
            written_documents.append(document)

        self._session.flush()
        return written_documents

    @router.post(
        "/projects/{project_id}/documents/excel",
        status_code=201,
        response_model=list[DocumentOutput],
        **kwargs,
    )
    def upload_documents_excel(
        self,
        project_id: int,
        upload_file: UploadFile,
        temp_file: NamedTemporaryFile = Depends(get_excel_temp_file),
    ) -> Iterator[Document]:
        return self._bulk_create_update_documents(
            project_id, self._process_upload(upload_file, temp_file)
        )
